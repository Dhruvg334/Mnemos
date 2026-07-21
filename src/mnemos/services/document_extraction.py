from __future__ import annotations

import csv
import io
from dataclasses import dataclass

MAX_EXTRACTED_CHARS = 2_000_000
MAX_PDF_PAGES = 500
MAX_SHEETS = 30
MAX_ROWS_PER_SHEET = 10_000
MAX_COLUMNS = 100
MAX_CELL_CHARS = 4_000
CHUNK_SIZE = 1_600
CHUNK_OVERLAP = 200


class ExtractionError(ValueError):
    code = "EXTRACTION_FAILED"


class ImageOnlyPdfError(ExtractionError):
    code = "IMAGE_ONLY_PDF"


@dataclass(frozen=True)
class ExtractedSection:
    locator: str
    text: str
    page_or_sheet: str | None = None
    section: str | None = None


def _bounded(value: str) -> str:
    return value.replace("\x00", "").strip()[:MAX_EXTRACTED_CHARS]


def extract_document(data: bytes, mime_type: str, filename: str) -> list[ExtractedSection]:
    lower = filename.lower()
    if mime_type == "application/pdf" or lower.endswith(".pdf"):
        return _extract_pdf(data)
    if mime_type in {"text/plain", "text/markdown", "text/x-markdown"} or lower.endswith(
        (".txt", ".md", ".markdown")
    ):
        return _extract_text(data)
    if mime_type in {"text/csv", "application/csv"} or lower.endswith(".csv"):
        return _extract_csv(data)
    if (
        mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        or lower.endswith(".docx")
    ):
        return _extract_docx(data)
    if (
        mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        or lower.endswith(".xlsx")
    ):
        return _extract_xlsx(data)
    raise ExtractionError("Unsupported document format.")


def _extract_text(data: bytes) -> list[ExtractedSection]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("utf-8", errors="replace")
    text = _bounded(text.replace("\r\n", "\n").replace("\r", "\n"))
    if not text:
        raise ExtractionError("No readable text was found.")
    return [ExtractedSection(locator="document", text=text)]


def _extract_pdf(data: bytes) -> list[ExtractedSection]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    sections: list[ExtractedSection] = []
    for index, page in enumerate(reader.pages[:MAX_PDF_PAGES], start=1):
        text = _bounded(page.extract_text() or "")
        if text:
            sections.append(
                ExtractedSection(locator=f"page:{index}", page_or_sheet=str(index), text=text)
            )
    if not sections:
        raise ImageOnlyPdfError(
            "No readable text was found. OCR is not currently available for image-only PDFs."
        )
    return sections


def _extract_csv(data: bytes) -> list[ExtractedSection]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for index, row in enumerate(reader):
        if index >= MAX_ROWS_PER_SHEET:
            break
        rows.append([cell[:MAX_CELL_CHARS] for cell in row[:MAX_COLUMNS]])
    if not rows:
        raise ExtractionError("The CSV contains no readable rows.")
    header = rows[0]
    rendered = [" | ".join(header)]
    for row_no, row in enumerate(rows[1:], start=2):
        rendered.append(
            f"Row {row_no}: "
            + " | ".join(
                f"{header[i] if i < len(header) else f'Column {i + 1}'}={value}"
                for i, value in enumerate(row)
            )
        )
    return [
        ExtractedSection(
            locator=f"rows:1-{len(rows)}", text=_bounded("\n".join(rendered)), page_or_sheet="CSV"
        )
    ]


def _extract_docx(data: bytes) -> list[ExtractedSection]:
    from docx import Document as DocxDocument

    doc = DocxDocument(io.BytesIO(data))
    sections: list[ExtractedSection] = []
    heading = None
    buffer: list[str] = []

    def flush() -> None:
        nonlocal buffer
        text = _bounded("\n".join(buffer))
        if text:
            sections.append(
                ExtractedSection(locator=heading or "document", section=heading, text=text)
            )
        buffer = []

    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue
        if paragraph.style and paragraph.style.name.lower().startswith("heading"):
            flush()
            heading = text
        else:
            buffer.append(text)
    flush()
    if not sections:
        raise ExtractionError("The DOCX contains no readable paragraphs.")
    return sections


def _extract_xlsx(data: bytes) -> list[ExtractedSection]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sections: list[ExtractedSection] = []
    for sheet in workbook.worksheets[:MAX_SHEETS]:
        lines = []
        for row_no, row in enumerate(
            sheet.iter_rows(max_row=MAX_ROWS_PER_SHEET, max_col=MAX_COLUMNS, values_only=True),
            start=1,
        ):
            values = ["" if value is None else str(value)[:MAX_CELL_CHARS] for value in row]
            if any(values):
                lines.append(f"Row {row_no}: " + " | ".join(values))
        text = _bounded("\n".join(lines))
        if text:
            sections.append(
                ExtractedSection(
                    locator=f"sheet:{sheet.title}", page_or_sheet=sheet.title, text=text
                )
            )
    if not sections:
        raise ExtractionError("The XLSX contains no readable cells.")
    return sections


def chunk_sections(sections: list[ExtractedSection]) -> list[tuple[ExtractedSection, str]]:
    chunks: list[tuple[ExtractedSection, str]] = []
    for section in sections:
        start = 0
        while start < len(section.text):
            end = min(len(section.text), start + CHUNK_SIZE)
            chunks.append((section, section.text[start:end].strip()))
            if end == len(section.text):
                break
            start = max(start + 1, end - CHUNK_OVERLAP)
    return [(section, text) for section, text in chunks if text]
